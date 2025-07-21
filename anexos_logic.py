import os
import unicodedata
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

def run_all_anexos(file_path, root_out):
    """
    Ejecuta los cinco anexos sobre `file_path` y va dejando los resultados
    en subcarpetas dentro de `root_out`.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # puedes adaptar aquí la lógica de elegir hoja si quieres
    hoja = wb.sheetnames[0]
    ws   = wb[hoja]

    header_row = detect_header(ws)
    raw_hdr    = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    norm_hdr   = [normalize(h) for h in raw_hdr]

    # Mapeo carpeta → función
    mapping = {
        "Anexo8_QR":     generate_anexo8,
        "Anexo9_Incorp": generate_anexo9,
        "Anexo10_Modif": generate_anexo10,
        "Anexo11_Desinc":generate_anexo11,
        "Anexo12_UbicCre":generate_anexo12,
    }

    for subdir_name, func in mapping.items():
        subdir = os.path.join(root_out, subdir_name)
        os.makedirs(subdir, exist_ok=True)
        func(ws, header_row, norm_hdr, subdir)
def normalize(text):
    """Convierte a mayúsculas sin tildes ni marcas diacríticas."""
    if text is None:
        return ""
    s = unicodedata.normalize("NFD", str(text).upper())
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")

def find_col(norm_headers, *keywords):
    """Devuelve el índice de la primera columna cuyo header contenga todas las keywords."""
    for i, nh in enumerate(norm_headers):
        if all(k in nh for k in keywords):
            return i
    raise ValueError(f"No encontré columna con {' & '.join(keywords)}")

def get_unique_filename(path):
    """Si el archivo existe, añade _1, _2, ... antes de la extensión."""
    base, ext = os.path.splitext(path)
    i = 1
    new = path
    while os.path.exists(new):
        new = f"{base}_{i}{ext}"
        i += 1
    return new

def get_unique_foldername(path):
    """Si la carpeta existe, añade _1, _2, ... al final."""
    i = 1
    new = path
    while os.path.exists(new):
        new = f"{path}_{i}"
        i += 1
    return new

def elegir_hoja(sheet_names):
    """Muestra un diálogo para elegir una hoja."""
    dlg = tk.Toplevel()
    dlg.title("Elige una hoja")
    dlg.geometry("300x300")
    dlg.grab_set()

    tk.Label(dlg, text="Selecciona la hoja:").pack(pady=5)
    lb = tk.Listbox(dlg)
    lb.pack(fill="both", expand=True, padx=10)
    for name in sheet_names:
        lb.insert("end", name)

    sel = {"hoja": None}
    def on_select():
        idx = lb.curselection()
        if idx:
            sel["hoja"] = lb.get(idx[0])
            dlg.destroy()
        else:
            messagebox.showwarning("Atención", "Debes seleccionar una hoja.")
    def on_cancel():
        dlg.destroy()

    btns = tk.Frame(dlg); btns.pack(pady=5)
    tk.Button(btns, text="Seleccionar", command=on_select, width=10).pack(side="left", padx=5)
    tk.Button(btns, text="Cancelar",  command=on_cancel, width=10).pack(side="right", padx=5)

    dlg.wait_window()
    return sel["hoja"]

def detect_header(ws):
    """Encuentra la fila que contiene DENOMINACION y NIVEL8."""
    for r in range(1, 21):
        vals = [normalize(c.value) for c in ws[r]]
        if "DENOMINACION" in vals and "NIVEL8" in vals:
            return r
    raise ValueError("No encontré fila de encabezado (buscando DENOMINACION y NIVEL8).")

def generate_anexo8(ws, header_row, norm_hdr, out_dir):
    """Anexo 8: pendientes de QR (filtra 'NO QR' en NIVEL7)."""
    C7 = find_col(norm_hdr, "NIVEL7")
    CAMPOS = {
        "Sitio":                        ["CAMPO","CLASIFICACION"],
        "Equipo":                       ["NIVEL8"],
        "Denominación de objeto técnico":["DENOMINACION"],
        "Tipo de equipo":               ["TIPO","EQUIPO"],
        "Tp.objeto técnico":            ["TP.OBJETO","TECNICO"],
        "Ubicación técnica":            ["UBICACION","TECNICA","SUPERIOR"],
        "No QR":                        ["NIVEL7"]
    }
    col_idx = {out: find_col(norm_hdr, *keys) for out, keys in CAMPOS.items()}

    rows = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=False):
        val = row[C7].value
        if isinstance(val, str) and val.strip().upper() == "NO QR":
            rec = {out: row[idx].value for out, idx in col_idx.items()}
            rows.append(rec)
    if not rows:
        return
    df = pd.DataFrame(rows)
    for sitio, grp in df.groupby("Sitio"):
        df_site = grp.drop(columns=["Sitio"])
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in str(sitio))
        path = get_unique_filename(os.path.join(out_dir, f"Anexo8_{safe}.xlsx"))
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            df_site.to_excel(writer, sheet_name="Anexo8", startrow=5, index=False)
            book  = writer.book; sheet = writer.sheets["Anexo8"]
            fmt_title    = book.add_format({"bold":True,"font_size":16,"align":"center","valign":"vcenter"})
            fmt_subtitle = book.add_format({"bold":True,"font_size":12,"align":"center","valign":"vcenter"})
            fmt_hdr      = book.add_format({"bold":True,"font_color":"#FFFFFF","bg_color":"#305496","border":1,"align":"center","valign":"vcenter"})
            fmt_cell     = book.add_format({"border":1,"valign":"vcenter"})
            n = len(df_site.columns)
            sheet.merge_range(0,0,2,n-1,f"INFORME DE ANÁLISIS DE CRITICIDAD A ESTACIÓN {sitio}", fmt_title)
            sheet.merge_range(3,0,3,n-1,"Anexo 8 – Listado de equipos pendientes por asignación de QR en campo", fmt_subtitle)
            sheet.set_row(0,30); sheet.set_row(1,20); sheet.set_row(2,20); sheet.set_row(3,25)
            for c, col in enumerate(df_site.columns):
                sheet.write(5,c,col,fmt_hdr)
                w = max(df_site[col].astype(str).map(len).max(), len(col)) + 2
                sheet.set_column(c,c,w,fmt_cell)

def generate_anexo9(ws, header_row, norm_hdr, out_dir):
    """Anexo 9: incorporación SAP (texto verde puro en NIVEL8)."""
    C8 = find_col(norm_hdr, "NIVEL8")
    GREEN = "FF00B050"
    CAMPOS = {
        "Identificación SAP":            ["COD","SAP"],
        "Equipo":                        ["NIVEL8"],
        "Denominación de objeto técnico":["DENOMINACION"],
        "Tipo de equipo":                ["TIPO","EQUIPO"],
        "Tp.objeto técnico":             ["TP.OBJETO","TECNICO"],
        "Peso bruto":                    ["PESO","BRUTO"],
        "Tamaño/Dimensión":              ["TAMAÑO"],
        "Número de inventario":          ["INVENTARIO"],
        "Fabricante del activo fijo":    ["FABRICANTE","ACTIVO"],
        "País de fabricación":           ["PAIS","FABRICACION"],
        "Denominación de tipo":          ["DENOMINACION","TIPO"],
        "Año de construcción":           ["ANO","CONSTRUCCION"],
        "Mes de construcción":           ["MES","CONSTRUCCION"],
        "Número de pieza de fabricante": ["NUMERO","PIEZA"],
        "Fabricante número de serie":    ["NUMERO","SERIE"],
        "Centro emplazamiento":          ["CENTRO","EMPLAZ"],
        "Emplazamiento":                 ["EMPLAZAMIENTO"],
        "Área de empresa":               ["AREA","EMPRESA"],
        "Indicador ABC":                 ["ASP"],
        "Campo de clasificación":        ["CAMPO","CLASIFICACION"],
        "Sociedad":                      ["SOCIEDAD"],
        "Centro de coste":               ["CENTRO","COSTE"],
        "Centro planificación":          ["CENTRO","PLANIF"],
        "Grupo planificación":           ["GRUPO","PLANIF"],
        "Pto.tbjo.responsable":          ["PTO.TBJO","RESPONSABLE"],
        "Perfil de catálogo":            ["PERFIL","CATALOGO"],
        "Ubicación técnica":             ["UBICACION","TECNICA","SUPERIOR"],
    }
    col_idx = {}
    for out, keys in CAMPOS.items():
        try:
            col_idx[out] = find_col(norm_hdr, *keys)
        except ValueError:
            pass

    registros = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=False):
        raw = getattr(row[C8].font.color, "rgb", "")
        rgb_str = str(raw) if raw is not None else ""
        if GREEN not in rgb_str.upper():
            continue
        rec = {out: row[idx].value for out, idx in col_idx.items()}
        registros.append(rec)
    if not registros:
        return

    df = pd.DataFrame(registros)
    cols = df.columns.tolist()
    if "Identificación SAP" in cols and "Denominación de objeto técnico" in cols:
        cols.remove("Identificación SAP")
        i = cols.index("Denominación de objeto técnico")
        cols.insert(i, "Identificación SAP")
        df = df[cols]

    for sitio, grp in df.groupby("Campo de clasificación"):
        df_site = grp.drop(columns=["Campo de clasificación"], errors="ignore")
        safe    = "".join(c if c.isalnum() or c in " _-" else "_" for c in str(sitio))
        outp    = get_unique_filename(os.path.join(out_dir, f"Anexo9_{safe}.xlsx"))
        with pd.ExcelWriter(outp, engine="xlsxwriter") as writer:
            df_site.to_excel(writer, sheet_name="Anexo9", startrow=5, index=False)
            book  = writer.book; sheet = writer.sheets["Anexo9"]
            fmt_title    = book.add_format({"bold":True,"font_size":16,"align":"center","valign":"vcenter"})
            fmt_subtitle = book.add_format({"bold":True,"font_size":12,"align":"center","valign":"vcenter"})
            fmt_hdr      = book.add_format({"bold":True,"font_color":"#FFFFFF","bg_color":"#305496","border":1,"align":"center","valign":"vcenter"})
            fmt_cell     = book.add_format({"border":1,"valign":"vcenter"})
            ncol = len(df_site.columns)
            sheet.merge_range(0,0,2,ncol-1,f"INFORME DE ANÁLISIS DE CRITICIDAD A ESTACIÓN {sitio}", fmt_title)
            sheet.merge_range(3,0,3,ncol-1,"Anexo 9 – Listado de equipos de incorporación en SAP (060)", fmt_subtitle)
            sheet.set_row(0,30); sheet.set_row(1,20); sheet.set_row(2,20); sheet.set_row(3,25)
            for c, col in enumerate(df_site.columns):
                sheet.write(5,c,col, fmt_hdr)
                w = max(df_site[col].astype(str).map(len).max(), len(col)) + 2
                sheet.set_column(c,c,w, fmt_cell)

def generate_anexo10(ws, header_row, norm_hdr, out_dir):
    """Anexo 10: equipos modificados (texto amarillo puro en NIVEL8)."""
    C8     = find_col(norm_hdr, "NIVEL8")
    YELLOW = "FFFF00"
    CAMPOS = {
        "Sitio":                         ["CAMPO","CLASIFICACION"],
        "Equipo":                        ["NIVEL8"],
        "Denominación de objeto técnico":["DENOMINACION"],
        "Ubicación técnica":             ["UBICACION","TECNICA","SUPERIOR"]
    }
    col_idx = {out: find_col(norm_hdr, *keys) for out, keys in CAMPOS.items()}

    rows = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=False):
        raw = getattr(row[C8].font.color, "rgb", "")
        rgb_str = str(raw) if raw is not None else ""
        if not rgb_str.upper().endswith(YELLOW):
            continue
        rec = {out: row[idx].value for out, idx in col_idx.items()}
        rows.append(rec)
    if not rows:
        return
    df = pd.DataFrame(rows)
    for sitio, grp in df.groupby("Sitio"):
        df_site = grp.drop(columns=["Sitio"])
        safe    = "".join(c if c.isalnum() or c in " _-" else "_" for c in str(sitio))
        path    = get_unique_filename(os.path.join(out_dir, f"Anexo10_{safe}.xlsx"))
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            df_site.to_excel(writer, sheet_name="Anexo10", startrow=5, index=False)
            book  = writer.book; sheet = writer.sheets["Anexo10"]
            fmt_title    = book.add_format({"bold":True,"font_size":16,"align":"center","valign":"vcenter"})
            fmt_subtitle = book.add_format({"bold":True,"font_size":12,"align":"center","valign":"vcenter"})
            fmt_hdr      = book.add_format({"bold":True,"font_color":"#FFFFFF","bg_color":"#305496","border":1,"align":"center","valign":"vcenter"})
            fmt_cell     = book.add_format({"border":1,"valign":"vcenter"})
            n = len(df_site.columns)
            sheet.merge_range(0,0,2,n-1,f"INFORME DE ANÁLISIS DE CRITICIDAD A ESTACIÓN {sitio}", fmt_title)
            sheet.merge_range(3,0,3,n-1,"Anexo 10 – Listado de equipos modificados en SAP", fmt_subtitle)
            sheet.set_row(0,30); sheet.set_row(1,20); sheet.set_row(2,20); sheet.set_row(3,25)
            for c, col in enumerate(df_site.columns):
                sheet.write(5,c,col, fmt_hdr)
                w = max(df_site[col].astype(str).map(len).max(), len(col)) + 2
                sheet.set_column(c,c,w, fmt_cell)

def generate_anexo11(ws, header_row, norm_hdr, out_dir):
    """Anexo 11: desincorporaciones (texto rojo puro en NIVEL8)."""
    C8  = find_col(norm_hdr, "NIVEL8")
    RED = "FF0000"
    CAMPOS = {
        "Sitio":                       ["CAMPO","CLASIFICACION"],
        "Identificación SAP":          ["COD","SAP"],
        "DENOMINACIÓN":                ["DENOMINACION"],
        "Tipo de equipo":              ["TIPO","EQUIPO"],
        "Tp.objeto técnico":           ["TP.OBJETO","TECNICO"],
        "Centro planif.":              ["CENTRO","PLANIF"],
        "Ubicación técnica superior":  ["UBICACION","TECNICA","SUPERIOR"]
    }
    col_idx = {out: find_col(norm_hdr, *keys) for out, keys in CAMPOS.items()}

    rows = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=False):
        raw = getattr(row[C8].font.color, "rgb", "")
        rgb_str = str(raw) if raw is not None else ""
        if not rgb_str.upper().endswith(RED):
            continue
        rec = {out: row[idx].value for out, idx in col_idx.items()}
        rows.append(rec)
    if not rows:
        return
    df = pd.DataFrame(rows)
    for sitio, grp in df.groupby("Sitio"):
        df_site = grp.copy()  # mantenemos Sitio
        safe    = "".join(c if c.isalnum() or c in " _-" else "_" for c in str(sitio))
        outp    = get_unique_filename(os.path.join(out_dir, f"Anexo11_{safe}.xlsx"))
        with pd.ExcelWriter(outp, engine="xlsxwriter") as writer:
            df_site.to_excel(writer, sheet_name="Anexo11", startrow=5, index=False)
            book  = writer.book; sheet = writer.sheets["Anexo11"]
            title_fmt    = book.add_format({"bold":True,"font_size":16,"align":"center","valign":"vcenter"})
            subtitle_fmt = book.add_format({"bold":True,"font_size":12,"align":"center","valign":"vcenter"})
            header_fmt   = book.add_format({"bold":True,"font_color":"#FFFFFF","bg_color":"#305496","border":1,"align":"center","valign":"vcenter"})
            cell_fmt     = book.add_format({"border":1,"valign":"vcenter"})
            n = df_site.shape[1]
            sheet.merge_range(0,0,2,n-1,f"INFORME DE ANÁLISIS DE CRITICIDAD A ESTACIÓN {sitio}", title_fmt)
            sheet.merge_range(3,0,3,n-1,"Anexo 11 – Listado de equipos desincorporados en SAP", subtitle_fmt)
            sheet.set_row(0,30); sheet.set_row(1,20); sheet.set_row(2,20); sheet.set_row(3,25)
            for c, col in enumerate(df_site.columns):
                sheet.write(5,c,col, header_fmt)
                w = max(df_site[col].astype(str).map(len).max(), len(col)) + 2
                sheet.set_column(c,c,w,cell_fmt)

def generate_anexo12(ws, header_row, norm_hdr, out_dir):
    """Anexo 12: ubicaciones técnicas creadas (verde en NIVEL5–7)."""
    LEVEL5  = find_col(norm_hdr, "NIVEL5")
    LEVEL6  = find_col(norm_hdr, "NIVEL6")
    LEVEL7  = find_col(norm_hdr, "NIVEL7")
    idx_site= find_col(norm_hdr, "CAMPO","CLASIFICACION")
    idx_tipo= find_col(norm_hdr, "TIPO","EQUIPO")
    idx_den = find_col(norm_hdr, "DENOMINACION")
    idx_tp  = find_col(norm_hdr, "TP.OBJETO","TECNICO")
    idx_ce  = find_col(norm_hdr, "CENTRO","EMPLAZ")
    idx_soc = find_col(norm_hdr, "SOCIEDAD")
    idx_cc  = find_col(norm_hdr, "CENTRO","COSTE")
    idx_cp  = find_col(norm_hdr, "CENTRO","PLANIF")
    idx_uts = find_col(norm_hdr, "UBICACION","TECNICA","SUPERIOR")

    GREEN = "FF00B050"
    registros = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=False):
        for idx in (LEVEL5, LEVEL6, LEVEL7):
            cell = row[idx]
            if not cell.value: continue
            fc = cell.font.color
            raw_font = getattr(fc, "rgb", "")
            rgb_font = str(raw_font) if raw_font is not None else ""
            fill = cell.fill
            fg   = fill.fgColor if fill and fill.patternType=="solid" else None
            raw_fill = getattr(fg, "rgb", "")
            rgb_fill = str(raw_fill) if raw_fill is not None else ""
            if GREEN in rgb_font.upper() or GREEN in rgb_fill.upper():
                registros.append({
                    "Sitio":      row[idx_site].value,
                    "Ubicación técnica": cell.value,
                    "Tipo ubic.técnica": row[idx_tipo].value,
                    "Denominación de la ubicación técnica": row[idx_den].value,
                    "Tp.objeto técnico": row[idx_tp].value,
                    "Centro emplazamiento": row[idx_ce].value,
                    "Campo de clasificación": row[idx_site].value,
                    "Sociedad":    row[idx_soc].value,
                    "Centro de coste": row[idx_cc].value,
                    "Centro planificación": row[idx_cp].value,
                    "Ubicación técnica superior": row[idx_uts].value
                })
                break

    if not registros:
        return
    df = pd.DataFrame(registros)
    df.index += 1
    df.insert(0, "#", df.index)
    cols = ["#", "Ubicación técnica", "Tipo ubic.técnica",
            "Denominación de la ubicación técnica", "Tp.objeto técnico",
            "Centro emplazamiento", "Campo de clasificación",
            "Sociedad", "Centro de coste", "Centro planificación",
            "Ubicación técnica superior"]
    df = df[cols + ["Sitio"]]

    for sitio, grp in df.groupby("Sitio"):
        df_site = grp.drop(columns=["Sitio"])
        safe    = "".join(c if c.isalnum() or c in " _-" else "_" for c in str(sitio))
        outp    = get_unique_filename(os.path.join(out_dir, f"Anexo12_{safe}.xlsx"))
        with pd.ExcelWriter(outp, engine="xlsxwriter") as writer:
            df_site.to_excel(writer, sheet_name="Anexo12", startrow=5, index=False)
            book = writer.book; sheet = writer.sheets["Anexo12"]
            title_fmt = book.add_format({"bold":True,"font_size":16,"align":"center","valign":"vcenter"})
            sub_fmt   = book.add_format({"bold":True,"font_size":12,"align":"center","valign":"vcenter"})
            hdr_fmt   = book.add_format({"bold":True,"font_color":"#FFFFFF","bg_color":"#305496","border":1,"align":"center","valign":"vcenter"})
            cell_fmt  = book.add_format({"border":1,"valign":"vcenter"})
            ncols = len(df_site.columns)
            sheet.merge_range(0,0,2,ncols-1,f"INFORME DE ANÁLISIS DE CRITICIDAD A ESTACIÓN {sitio}", title_fmt)
            sheet.merge_range(3,0,3,ncols-1,"Anexo 12 – Listado de ubicaciones técnicas creadas en SAP", sub_fmt)
            sheet.set_row(0,30);sheet.set_row(1,20);sheet.set_row(2,20);sheet.set_row(3,25)
            for i, col in enumerate(df_site.columns):
                sheet.write(5,i,col, hdr_fmt)
                w = max(df_site[col].astype(str).map(len).max(), len(col)) + 2
                sheet.set_column(i,i,w, cell_fmt)

def main():
    tk.Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="Seleccione el archivo Árbol de Equipos",
        filetypes=[("Excel","*.xlsx")]
    )
    if not file_path:
        return

    wb     = openpyxl.load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    hoja   = elegir_hoja(sheets) if len(sheets) > 1 else sheets[0]
    if not hoja:
        return
    ws = wb[hoja]

    header_row = detect_header(ws)
    raw_hdr    = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    norm_hdr   = [normalize(h) for h in raw_hdr]

    # Carpeta raíz
    base     = os.path.join(os.path.dirname(file_path), "Anexos_Todos")
    root_out = get_unique_foldername(base)
    os.makedirs(root_out)

    # Ejecutar cada anexo
    mapping = {
        8:  ("Anexo8_QR",     generate_anexo8),
        9:  ("Anexo9_Incorp", generate_anexo9),
        10: ("Anexo10_Modif", generate_anexo10),
        11: ("Anexo11_Desinc",generate_anexo11),
        12: ("Anexo12_UbicCre",generate_anexo12),
    }
    for num, (sub, func) in mapping.items():
        subdir = os.path.join(root_out, sub)
        os.makedirs(subdir, exist_ok=True)
        func(ws, header_row, norm_hdr, subdir)

    messagebox.showinfo("¡Listo!", f"Se generaron todos los Anexos en:\n{root_out}")

if __name__ == "__main__":
    main()
